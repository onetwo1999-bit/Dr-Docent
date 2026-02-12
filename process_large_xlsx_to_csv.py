"""
엑셀 → Supabase 업로드용 CSV 변환 (메모리 절약 청크 처리)

- openpyxl read_only=True로 엑셀 스트리밍, 1만 행 단위 청킹 후 CSV에 이어붙이기
- 9개 영문 컬럼, 1인분 영양 환산, clinical_insight, synthetic_qa 생성 (동일 로직)

실행 예:
  python process_large_xlsx_to_csv.py
    → raw_food_db.xlsx → processed_food_db_final_250k.csv (기본 25만 건용)

  python process_large_xlsx_to_csv.py raw_food_db_part2.xlsx processed_food_db_part2.csv
    → part2 전용 (기존 _final_250k.csv 파일은 건드리지 않음)
"""

import json
import sys
from pathlib import Path

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("❌ openpyxl 필요: pip install openpyxl")

# 입출력: 인자 2개 주면 part2 등 별도 파일로 동작, 없으면 기본 25만 건용
if len(sys.argv) >= 3:
    INPUT_XLSX = sys.argv[1]
    OUTPUT_CSV = sys.argv[2]
else:
    INPUT_XLSX = "raw_food_db.xlsx"
    OUTPUT_CSV = "processed_food_db_final_250k.csv"

CHUNK_SIZE = 10_000

NUTRIENT_COLS = ["calories", "protein", "fat", "carbs", "sugar", "sodium"]
FINAL_COLS = [
    "food_name", "calories", "protein", "fat", "carbs", "sugar", "sodium",
    "clinical_insight", "synthetic_qa",
]

# 원본 한글 컬럼 → 수파베이스 영문 컬럼 (여러 표기 허용)
COLUMN_MAPPING = [
    ("식품명", "food_name"),
    ("에너지(kcal)", "calories"),
    ("에너지", "calories"),
    ("단백질(g)", "protein"),
    ("단백질", "protein"),
    ("지방(g)", "fat"),
    ("지방", "fat"),
    ("탄수화물(g)", "carbs"),
    ("탄수화물", "carbs"),
    ("당류(g)", "sugar"),
    ("당류", "sugar"),
    ("나트륨(mg)", "sodium"),
    ("나트륨", "sodium"),
]

WEIGHT_KEYWORDS = ["1회제공량", "1회 제공량", "총내용량", "중량", "내용량", "1회분량"]


def find_weight_column(columns):
    for kw in WEIGHT_KEYWORDS:
        for c in columns:
            if kw in str(c):
                return c
    return None


def apply_column_mapping(df: pd.DataFrame) -> pd.DataFrame:
    renames = {}
    seen_eng = set()
    for kor, eng in COLUMN_MAPPING:
        if kor in df.columns and eng not in seen_eng:
            renames[kor] = eng
            seen_eng.add(eng)
    return df.rename(columns=renames)


def apply_serving_conversion(df: pd.DataFrame, weight_col) -> pd.DataFrame:
    df = df.copy()
    if weight_col is None or weight_col not in df.columns:
        for col in NUTRIENT_COLS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df
    weight_series = pd.to_numeric(df[weight_col], errors="coerce").fillna(100)
    ratio = weight_series / 100.0
    for col in NUTRIENT_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            df[col] = (df[col] * ratio).round(1)
    return df


def add_clinical_insight(row: pd.Series) -> str:
    sodium = row.get("sodium") or 0
    sugar = row.get("sugar") or 0
    protein = row.get("protein") or 0
    parts = []
    if sodium >= 500:
        parts.append(f"1인분 기준 나트륨 {int(sodium)}mg로 고혈압 주의가 필요합니다.")
    if sugar >= 10:
        parts.append(f"당류 {sugar}g 포함으로 당뇨 관리 시 양을 조절하세요.")
    if protein >= 15:
        parts.append("단백질이 풍부해 근성장·유지에 도움이 됩니다.")
    if not parts:
        return "균형 잡힌 영양 성분입니다."
    return " ".join(parts)


def add_synthetic_qa(row: pd.Series) -> str:
    name = row.get("food_name") or ""
    cal = row.get("calories") or 0
    insight = row.get("clinical_insight") or ""
    return json.dumps({
        "question": f"{name}의 칼로리와 영양은?",
        "answer": f"1인분 기준 {cal}kcal이며, {insight}"
    }, ensure_ascii=False)


def stream_xlsx_rows(path: Path, chunk_size: int):
    """엑셀을 read_only로 열고 행을 청크 단위로 yield (메모리 절약)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = None
    chunk = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c) if c is not None else "" for c in row]
            continue
        chunk.append(list(row))
        if len(chunk) >= chunk_size:
            yield header, chunk
            chunk = []
    if chunk:
        yield header, chunk
    wb.close()


def main():
    folder = Path(__file__).resolve().parent
    xlsx_path = folder / INPUT_XLSX
    csv_path = folder / OUTPUT_CSV

    if not xlsx_path.exists():
        print(f"❌ 현재 폴더에 {INPUT_XLSX} 이(가) 없습니다.")
        return

    print(f"📥 {INPUT_XLSX} → {OUTPUT_CSV} (청크당 {CHUNK_SIZE:,}행)")
    weight_col_global = None
    total_rows = 0
    first_chunk = True

    for header, rows in stream_xlsx_rows(xlsx_path, CHUNK_SIZE):
        n_cols = len(header)
        # 셀 개수가 헤더와 다를 수 있으므로 길이 맞춤
        normalized = [(r + [None] * n_cols)[:n_cols] for r in rows]
        df = pd.DataFrame(normalized, columns=header)
        # 첫 청크에서만 기준 중량 컬럼 결정
        if weight_col_global is None:
            weight_col_global = find_weight_column(df.columns)
            if weight_col_global:
                print(f"⚖️ 기준 중량 컬럼: '{weight_col_global}' (1인분 환산)")
            else:
                print("⚖️ 기준 중량 컬럼 없음 → 100g 기준 유지")

        df = apply_column_mapping(df)
        missing = [c for c in ["food_name"] + NUTRIENT_COLS if c not in df.columns]
        if missing:
            print(f"❌ 원본에 필수 컬럼 없음: {missing}. 원본 컬럼: {list(df.columns)[:20]}...")
            return

        df = apply_serving_conversion(df, weight_col_global)
        df["clinical_insight"] = df.apply(add_clinical_insight, axis=1)
        df["synthetic_qa"] = df.apply(add_synthetic_qa, axis=1)
        out = df[FINAL_COLS]

        out.to_csv(
            csv_path,
            index=False,
            encoding="utf-8-sig",
            mode="w" if first_chunk else "a",
            header=first_chunk,
        )
        total_rows += len(out)
        first_chunk = False

        # 1만 건마다 "X만 건 변환 완료" 로그 출력
        if total_rows and total_rows % 10_000 == 0:
            print(f"   ✅ {total_rows // 10_000}만 건 변환 완료")

    print(f"✅ {OUTPUT_CSV} 저장 완료 (총 {total_rows:,}행, 9개 컬럼)")


if __name__ == "__main__":
    main()
